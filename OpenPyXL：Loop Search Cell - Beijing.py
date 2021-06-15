import openpyxl
import jieba

# *** Province Filter *** #
province_wb = openpyxl.load_workbook("province_city_district_data/provinces.xlsx")
print(province_wb.sheetnames)
province_sheet = province_wb["Sheet1"]
province_B_columns = province_sheet["B"]
province_filter_list = []

for province_cell in province_B_columns:
    province_name = province_cell.value
    province_length = len(province_name)
    province_filter_list.append(province_name)
    if "自治区" in province_name:
        province_filter_list.append(province_name[:province_length - 3])
    else:
        province_filter_list.append(province_name[:province_length - 1])

print(province_filter_list)

# *** City Filter *** #
city_wb = openpyxl.load_workbook("province_city_district_data/cities.xlsx")
print(city_wb.sheetnames)
city_sheet = city_wb["Sheet1"]
city_B_columns = city_sheet["B"]
city_filter_list = []

for city_cell in city_B_columns:
    city_name = city_cell.value
    city_length = len(city_name)
    city_filter_list.append(city_name)
    if "自治州" in city_name:
        city_filter_list.append(city_name[:city_length - 3])
    elif "地区" in city_name:
        city_filter_list.append(city_name[:city_length - 2])
    else:
        city_filter_list.append(city_name[:city_length - 1])

print(city_filter_list)

city_wb = openpyxl.load_workbook("province_city_district_data/cities.xlsx")
print(city_wb.sheetnames)
city_sheet = city_wb["Sheet1"]
city_B_columns = city_sheet["B"]
city_filter_list = []

for city_cell in city_B_columns:
    city_name = city_cell.value
    city_length = len(city_name)
    city_filter_list.append(city_name)
    if "自治州" in city_name:
        city_filter_list.append(city_name[:city_length - 3])
    elif "地区" in city_name:
        city_filter_list.append(city_name[:city_length - 2])
    else:
        city_filter_list.append(city_name[:city_length - 1])

print(city_filter_list)

# *** Area Filter *** #
area_wb = openpyxl.load_workbook("province_city_district_data/areas.xlsx")
print(area_wb.sheetnames)
area_sheet = area_wb["Sheet1"]
area_B_columns = area_sheet["B"]
area_filter_list = []

for area_cell in area_B_columns:
    area_name = area_cell.value
    area_length = len(area_name)
    area_filter_list.append(area_name)
    if "开发区" in area_name:
        pass
    elif "示范区" in area_name:
        pass
    elif "管理区" in area_name:
        pass
    elif "园区" in area_name:
        pass
    elif "自治县" in area_name:
        area_filter_list.append(area_name[:area_length - 3])
    else:
        area_filter_list.append(area_name[:area_length - 1])

print(area_filter_list)

# *** Custom Filter *** #
custom_filter_list = ["首都", "自治", "自治区", "高新", "高新区", "天府", "新区", "天府新区", "浆洗街", "亚运村",
                      "人民", "政府", "西藏自治区人民政府", "驻", "办事处", "科", "附属", "国际",
                      "综合", "开放", "医疗", "医科", "医美", "医生", "医学美容", "医学", "医学院", "医学中心",
                      "医院", "美容", "美容外科", "整形外科", "眼科医院", "妇产医院", "口腔医院", "皮肤",
                      "盆底", "康复", "中心", "诊所", "门诊", "门诊部", "综合", "整形", "中西医", "中医", "西医", "结合", "特色",
                      "有限公司", "有限责任", "企业", "投资", "管理", "咨询", "普通", "合伙", "公司", "集团",
                      "科技", "开发", "科技开发",
                      "大学", "复旦", "复旦大学", "交通", "交通大学", "上海交通大学", "中医药大学", "首都医科大学", "四川大学",
                      "（", "长宁店", "）", "(", "奥克斯", "广场", "店", ")"]
print(custom_filter_list)

filter_list = province_filter_list + city_filter_list + area_filter_list + custom_filter_list
print(filter_list)

# *** Load Hospital Data *** #
# wb1 = openpyxl.load_workbook("data/Generated - Shanghai Hospital List.xlsx")
wb1 = openpyxl.load_workbook("data/Generated - Beijing Hospital List.xlsx")
# wb1 = openpyxl.load_workbook("data/Generated - Chengdu Hospital List.xlsx")
print(wb1.sheetnames)
sheet1 = wb1["Sheet1"]
Galderma_hospital_B_columns = sheet1["B"]

wb2 = openpyxl.load_workbook("data/Data Source - SoYoung List.xlsx")
print(wb2.sheetnames)
sheet2 = wb2["Sheet1"]
SoYoung_B_columns = sheet2["B"]

wb3 = openpyxl.load_workbook("data/Data Source - Allergan List.xlsx")
print(wb3.sheetnames)
sheet3 = wb3["Sheet1"]
Allergan_B_columns = sheet3["B"]

wb4 = openpyxl.load_workbook("data/Data Source - SAIC List.xlsx")
print(wb4.sheetnames)
sheet4 = wb4["Sheet1"]
SAIC_F_columns = sheet4["F"]

# City:
city_shanghai = "上海"
city_beijing = "北京"
city_chengdu = "成都"

def hospital_mapping(city):
    Galderma_hospital_row_index = 0

    for Galderma_Hospital_cell in Galderma_hospital_B_columns:
        keyword_list = []
        Galderma_hospital_row_index = Galderma_hospital_row_index + 1
        # 数据清洗："None"，"-"，"VLOOKUP"
        if Galderma_Hospital_cell.value is not None and Galderma_Hospital_cell.value != "-" and "VLOOKUP" not in Galderma_Hospital_cell.value:
            print(Galderma_Hospital_cell.value)

            # *** 1st Keyword Location *** #
            word_list = jieba.lcut(Galderma_Hospital_cell.value)
            print(", ".join(word_list))

            for word in word_list:
                if word not in filter_list:
                    keyword_list.append(word)

            keyword = "".join(keyword_list)
            print("Keyword：" + keyword)

            # *** 1st SoYoung Mapping *** #
            keyword_search_SoYoung_row_index = 0
            keyword_search_SoYoung_cell_value_list = []

            for SoYoung_cell in SoYoung_B_columns:
                keyword_search_SoYoung_row_index = keyword_search_SoYoung_row_index + 1
                # 数据清洗："None"，"-"，"VLOOKUP"
                if SoYoung_cell.value is not None and SoYoung_cell.value != "-" and "VLOOKUP" not in SoYoung_cell.value:
                    if city in str(sheet2['E' + str(keyword_search_SoYoung_row_index)].value):
                        if keyword != "" and keyword in SoYoung_cell.value:
                            print("SoYoung keyword mapping row index: " + str(keyword_search_SoYoung_row_index))
                            print("SoYoung keyword mapping cell value：" + str(SoYoung_cell.value))
                            keyword_search_SoYoung_cell_value_list.append(SoYoung_cell.value)

            keyword_SoYoung_mapping_list = keyword_search_SoYoung_cell_value_list
            keyword_SoYoung_mapping_string = ", ".join(keyword_SoYoung_mapping_list)

            print("1st SoYoung Mapping: " + keyword_SoYoung_mapping_string)
            print()

            # *** 2nd SoYoung Mapping *** #
            if keyword_SoYoung_mapping_string == "" and len(keyword) > 3:
                sub_keyword_list = []

                for i in range(0, len(keyword), 2):
                    sub_keyword_list.append(keyword[i:i + 2])

                print("SoYoung sub keyword list:" + str(sub_keyword_list))

                for sub_keyword in sub_keyword_list:
                    if len(sub_keyword) < 2:
                        break

                    sub_keyword_search_SoYoung_row_index = 0
                    sub_keyword_search_SoYoung_cell_value_list = []

                    if sub_keyword not in filter_list:
                        print("SoYoung sub keyword：" + sub_keyword)
                        for SoYoung_cell in SoYoung_B_columns:
                            sub_keyword_search_SoYoung_row_index = sub_keyword_search_SoYoung_row_index + 1
                            # 数据清洗："None"，"-"，"VLOOKUP"
                            if SoYoung_cell.value is not None and SoYoung_cell.value != "-" and "VLOOKUP" not in SoYoung_cell.value:
                                if city in str(sheet2['E' + str(sub_keyword_search_SoYoung_row_index)].value):
                                    if sub_keyword != "" and sub_keyword in SoYoung_cell.value:
                                        print("SoYoung sub-keyword mapping row index: " + str(
                                            sub_keyword_search_SoYoung_row_index))
                                        print("SoYoung sub-keyword mapping cell value：" + str(SoYoung_cell.value))
                                        sub_keyword_search_SoYoung_cell_value_list.append(SoYoung_cell.value)
                                        break

                    if sub_keyword_search_SoYoung_cell_value_list:
                        break

                keyword_SoYoung_mapping_list = sub_keyword_search_SoYoung_cell_value_list
                keyword_SoYoung_mapping_string = ", ".join(keyword_SoYoung_mapping_list)
                print("2nd SoYoung Mapping: " + keyword_SoYoung_mapping_string)
                print()

            # *** 1st Allergan Mapping *** #
            keyword_search_Allergan_row_index = 0
            keyword_search_Allergan_cell_value_list = []

            for Allergan_cell in Allergan_B_columns:
                keyword_search_Allergan_row_index = keyword_search_Allergan_row_index + 1
                # 数据清洗："None"，"-"，"VLOOKUP"
                if Allergan_cell.value is not None and Allergan_cell.value != "-" and "VLOOKUP" not in Allergan_cell.value:
                    if city in str(sheet3['M' + str(keyword_search_Allergan_row_index)].value):
                        if keyword != "" and keyword in Allergan_cell.value:
                            print("Allergan keyword mapping row index: " + str(keyword_search_Allergan_row_index))
                            print("Allergan keyword mapping cell value：" + str(Allergan_cell.value))
                            keyword_search_Allergan_cell_value_list.append(Allergan_cell.value)
                            break

            keyword_Allergan_mapping_list = keyword_search_Allergan_cell_value_list
            keyword_Allergan_mapping_string = ", ".join(keyword_Allergan_mapping_list)
            print("1st Allergan Mapping: " + keyword_Allergan_mapping_string)
            print()

            # *** 2nd Allergan Mapping *** #
            if keyword_Allergan_mapping_string == "" and len(keyword) > 3:
                sub_keyword_list = []

                for i in range(0, len(keyword), 2):
                    sub_keyword_list.append(keyword[i:i + 2])

                print("Allergan sub keyword list:" + str(sub_keyword_list))

                for sub_keyword in sub_keyword_list:
                    if len(sub_keyword) < 2:
                        break

                    sub_keyword_search_Allergan_row_index = 0
                    sub_keyword_search_Allergan_cell_value_list = []

                    if sub_keyword not in filter_list:
                        print("Allergan sub keyword：" + sub_keyword)

                        for Allergan_cell in Allergan_B_columns:
                            sub_keyword_search_Allergan_row_index = sub_keyword_search_Allergan_row_index + 1
                            # 数据清洗："None"，"-"，"VLOOKUP"
                            if Allergan_cell.value is not None and Allergan_cell.value != "-" and "VLOOKUP" not in Allergan_cell.value:
                                if city in str(sheet3['M' + str(sub_keyword_search_Allergan_row_index)].value):
                                    if sub_keyword != "" and sub_keyword in Allergan_cell.value:
                                        print("Allergan sub-keyword mapping row index: " + str(sub_keyword_search_Allergan_row_index))
                                        print("Allergan sub-keyword mapping cell value：" + str(Allergan_cell.value))
                                        sub_keyword_search_Allergan_cell_value_list.append(Allergan_cell.value)
                                        break

                    if sub_keyword_search_Allergan_cell_value_list:
                        break

                keyword_Allergan_mapping_list = sub_keyword_search_Allergan_cell_value_list
                keyword_Allergan_mapping_string = ", ".join(keyword_Allergan_mapping_list)
                print("2nd Allergan Mapping: " + keyword_Allergan_mapping_string)
                print()

            # *** 1st SAIC Mapping *** #
            keyword_search_SAIC_row_index = 0
            keyword_search_SAIC_cell_value_list = []

            for SAIC_cell in SAIC_F_columns:
                keyword_search_SAIC_row_index = keyword_search_SAIC_row_index + 1
                # 数据清洗："None"，"-"，"VLOOKUP"
                if SAIC_cell.value is not None and SAIC_cell.value != "-" and "VLOOKUP" not in SAIC_cell.value:
                    if city in str(sheet4['C' + str(keyword_search_SAIC_row_index)].value):
                        if keyword != "" and keyword in SAIC_cell.value:
                            print("SAIC keyword mapping row index: " + str(keyword_search_SAIC_row_index))
                            print("SAIC keyword mapping cell value：" + str(SAIC_cell.value))
                            keyword_search_SAIC_cell_value_list.append(SAIC_cell.value)
                            break

            keyword_SAIC_mapping_list = keyword_search_SAIC_cell_value_list
            keyword_SAIC_mapping_string = ", ".join(keyword_SAIC_mapping_list)
            print("1st SAIC Mapping: " + keyword_SAIC_mapping_string)
            print()

            # *** 2nd SAIC Mapping *** #
            if keyword_SAIC_mapping_string == "" and len(keyword) > 3:
                sub_keyword_list = []

                for i in range(0, len(keyword), 2):
                    sub_keyword_list.append(keyword[i:i + 2])

                print("SAIC sub keyword list:" + str(sub_keyword_list))

                for sub_keyword in sub_keyword_list:
                    if len(sub_keyword) < 2:
                        break

                    sub_keyword_search_SAIC_row_index = 0
                    sub_keyword_search_SAIC_cell_value_list = []

                    if sub_keyword not in filter_list:
                        print("SAIC sub keyword：" + sub_keyword)

                        for SAIC_cell in SAIC_F_columns:
                            sub_keyword_search_SAIC_row_index = sub_keyword_search_SAIC_row_index + 1
                            # 数据清洗："None"，"-"，"VLOOKUP"
                            if SAIC_cell.value is not None and SAIC_cell.value != "-" and "VLOOKUP" not in SAIC_cell.value:
                                if city in str(sheet4['C' + str(sub_keyword_search_SAIC_row_index)].value):
                                    if sub_keyword != "" and sub_keyword in SAIC_cell.value:
                                        print("SAIC sub-keyword mapping row index: " + str(
                                            sub_keyword_search_SAIC_row_index))
                                        print("SAIC sub-keyword mapping cell value：" + str(SAIC_cell.value))
                                        sub_keyword_search_SAIC_cell_value_list.append(SAIC_cell.value)
                                        break

                    if sub_keyword_search_SAIC_cell_value_list:
                        break

                keyword_SAIC_mapping_list = sub_keyword_search_SAIC_cell_value_list
                keyword_SAIC_mapping_string = ", ".join(keyword_SAIC_mapping_list)
                print("2nd SAIC Mapping: " + keyword_SAIC_mapping_string)
                print()

            sheet1["H" + str(Galderma_hospital_row_index)] = keyword_SoYoung_mapping_string
            sheet1["L" + str(Galderma_hospital_row_index)] = keyword_Allergan_mapping_string
            sheet1["M" + str(Galderma_hospital_row_index)] = keyword_SAIC_mapping_string


hospital_mapping(city_beijing)

sheet1["H1"] = "SoYoung 2400 非工商"
sheet1["L1"] = "Allergan 3270 非工商 （Allergan有Gal没有的，可以考虑为D/new, esp. toxin), May-18"
sheet1["M1"] = "SAIC 13K, May-13. 先vlookup 13K，再对USCC"

# wb1.save("data/Generated - Shanghai Hospital List.xlsx")
wb1.save("data/Generated - Beijing Hospital List.xlsx")
# wb1.save("data/Generated - Chengdu Hospital List.xlsx")
